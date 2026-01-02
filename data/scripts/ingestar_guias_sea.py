#!/usr/bin/env python3
"""
Script para ingestar guías y criterios de evaluación del SEA al corpus RAG.

Extrae las guías del archivo Excel con índice oficial del SEA,
descarga los PDFs y los procesa para el sistema RAG.
"""

import asyncio
import os
import sys
import json
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Tuple
import logging
import hashlib

# Agregar el directorio backend al path
# En Docker el backend está en /app, en local en ../../backend
if Path("/app").exists() and Path("/app/app").exists():
    sys.path.insert(0, "/app")
else:
    sys.path.insert(0, str(Path(__file__).parent.parent.parent / "backend"))

import httpx
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.ext.asyncio import async_sessionmaker
from sqlalchemy import text

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configuración
# Cuando se ejecuta en Docker, usar /app/data; en local usar ../data
DATA_BASE = Path("/app/data") if Path("/app/data").exists() else Path(__file__).parent.parent
EXCEL_PATH = DATA_BASE / "legal" / "índice de guías y DT_23-12-2025.xlsx"
DOWNLOAD_DIR = DATA_BASE / "legal" / "guias_sea"
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://mineria:mineria_dev_2024@db:5432/mineria"
).replace("postgresql://", "postgresql+asyncpg://")

# Mapeo de tipos de guía a categorías del sistema
TIPO_CATEGORIA = {
    "Guías para la aplicación de normas y reglamentos": "guia_normas",
    "Guías de descripción de proyectos": "guia_proyectos",
    "Guías relacionadas al artículo 11 de la Ley N°19.300": "guia_art11",
    "Guía sobre metodologías y modelos": "guia_metodologia",
    "Guías para la descripción del área de influencia": "guia_area_influencia",
    "Guías sobre Participación Ciudadana (PAC)": "guia_pac",
    "Guías para los Permisos Ambientales Sectoriales (PAS)": "guia_pas",
    "Criterio de evaluación en el SEIA": "criterio_evaluacion",
}

# Triggers Art. 11 detectables por keywords
TRIGGERS_KEYWORDS = {
    'a': ['salud', 'población', 'riesgo sanitario', 'contaminante', 'emisión'],
    'b': ['recurso natural', 'renovable', 'agua', 'suelo', 'flora', 'fauna', 'ecosistema'],
    'c': ['reasentamiento', 'comunidad humana', 'desplazamiento', 'relocalización'],
    'd': ['área protegida', 'glaciar', 'parque nacional', 'reserva', 'santuario', 'humedal', 'SNASPE'],
    'e': ['patrimonio', 'arqueológico', 'monumento', 'histórico', 'cultural'],
    'f': ['paisaje', 'turismo', 'valor turístico', 'belleza escénica'],
}


def extraer_guias_excel(xlsx_path: Path) -> Tuple[List[Dict], List[Dict]]:
    """
    Extrae guías y criterios del archivo Excel con sus URLs.

    Returns:
        Tupla (guías, criterios)
    """
    logger.info(f"Extrayendo datos de {xlsx_path}")

    def extraer_hyperlinks(zip_file, sheet_rels_path: str) -> Dict[str, str]:
        """Extrae mapeo rId -> URL desde el archivo de relaciones."""
        try:
            content = zip_file.read(sheet_rels_path).decode('utf-8')
            root = ET.fromstring(content)
            ns = '{http://schemas.openxmlformats.org/package/2006/relationships}'

            links = {}
            for rel in root.findall(f'.//{ns}Relationship'):
                rid = rel.get('Id')
                target = rel.get('Target', '')
                rel_type = rel.get('Type', '')
                if 'hyperlink' in rel_type.lower():
                    links[rid] = target
            return links
        except Exception as e:
            logger.warning(f"Error extrayendo hyperlinks de {sheet_rels_path}: {e}")
            return {}

    def extraer_refs_hyperlinks(zip_file, sheet_path: str) -> Dict[str, str]:
        """Extrae mapeo celda -> rId desde el sheet XML."""
        try:
            content = zip_file.read(sheet_path).decode('utf-8')
            root = ET.fromstring(content)
            ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'

            refs = {}
            for hl in root.findall(f'.//{ns}hyperlink'):
                ref = hl.get('ref')
                rid = hl.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                if ref and rid:
                    refs[ref] = rid
            return refs
        except Exception as e:
            logger.warning(f"Error extrayendo refs de {sheet_path}: {e}")
            return {}

    # Extraer hyperlinks del ZIP
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        links_guias = extraer_hyperlinks(z, 'xl/worksheets/_rels/sheet2.xml.rels')
        refs_guias = extraer_refs_hyperlinks(z, 'xl/worksheets/sheet2.xml')
        links_criterios = extraer_hyperlinks(z, 'xl/worksheets/_rels/sheet3.xml.rels')
        refs_criterios = extraer_refs_hyperlinks(z, 'xl/worksheets/sheet3.xml')

    # Cargar workbook
    wb = load_workbook(xlsx_path)

    # Procesar Guías
    guias = []
    sheet = wb['Guías']
    for row_num in range(8, sheet.max_row + 1):
        nombre = sheet.cell(row=row_num, column=4).value
        if not nombre:
            continue

        tipo = sheet.cell(row=row_num, column=2).value
        anio = sheet.cell(row=row_num, column=3).value
        resolucion = sheet.cell(row=row_num, column=5).value
        vigencia = sheet.cell(row=row_num, column=6).value

        # Buscar URL
        cell_ref = f'D{row_num}'
        url = None
        if cell_ref in refs_guias:
            rid = refs_guias[cell_ref]
            url = links_guias.get(rid)

        if url and vigencia and 'vigente' in str(vigencia).lower():
            guias.append({
                'nombre': str(nombre).strip(),
                'tipo': str(tipo).strip() if tipo else 'Guía SEA',
                'anio': str(anio) if anio else '',
                'resolucion': str(resolucion)[:200] if resolucion else '',
                'vigencia': str(vigencia) if vigencia else '',
                'url': url,
                'documento_tipo': 'Guía SEA',
            })

    # Procesar Criterios
    criterios = []
    sheet = wb['Criterios de Evaluación']
    for row_num in range(7, sheet.max_row + 1):
        nombre = sheet.cell(row=row_num, column=4).value
        if not nombre:
            continue

        tipo = sheet.cell(row=row_num, column=2).value
        anio = sheet.cell(row=row_num, column=3).value
        resolucion = sheet.cell(row=row_num, column=5).value
        vigencia = sheet.cell(row=row_num, column=6).value

        cell_ref = f'D{row_num}'
        url = None
        if cell_ref in refs_criterios:
            rid = refs_criterios[cell_ref]
            url = links_criterios.get(rid)

        if url and vigencia and 'vigente' in str(vigencia).lower():
            criterios.append({
                'nombre': str(nombre).strip(),
                'tipo': str(tipo).strip() if tipo else 'Criterio SEA',
                'anio': str(anio) if anio else '',
                'resolucion': str(resolucion)[:200] if resolucion else '',
                'vigencia': str(vigencia) if vigencia else '',
                'url': url,
                'documento_tipo': 'Criterio SEA',
            })

    logger.info(f"Extraídas {len(guias)} guías y {len(criterios)} criterios vigentes")
    return guias, criterios


async def descargar_pdf(url: str, destino: Path, client: httpx.AsyncClient) -> bool:
    """Descarga un PDF desde una URL."""
    try:
        # Normalizar URL
        if not url.startswith('http'):
            url = 'https://' + url
        url = url.replace('sea.gob.cl', 'www.sea.gob.cl')

        response = await client.get(url, follow_redirects=True, timeout=60.0)
        response.raise_for_status()

        destino.parent.mkdir(parents=True, exist_ok=True)
        destino.write_bytes(response.content)

        logger.debug(f"Descargado: {destino.name}")
        return True

    except Exception as e:
        logger.warning(f"Error descargando {url}: {e}")
        return False


def extraer_texto_pdf(ruta_pdf: Path) -> str:
    """Extrae texto de un PDF usando PyMuPDF."""
    try:
        import fitz  # PyMuPDF

        texto = ""
        with fitz.open(str(ruta_pdf)) as doc:
            for pagina in doc:
                texto += pagina.get_text()

        return texto.strip()
    except Exception as e:
        logger.error(f"Error extrayendo texto de {ruta_pdf}: {e}")
        return ""


def calcular_sha256(ruta_archivo: Path) -> str:
    """Calcula hash SHA256 del contenido del archivo."""
    sha256 = hashlib.sha256()
    with open(ruta_archivo, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            sha256.update(chunk)
    return sha256.hexdigest()


def contar_paginas_pdf(ruta_pdf: Path) -> int:
    """Cuenta las páginas de un PDF usando PyMuPDF."""
    try:
        import fitz
        with fitz.open(str(ruta_pdf)) as doc:
            return len(doc)
    except Exception as e:
        logger.warning(f"Error contando páginas de {ruta_pdf}: {e}")
        return 0


async def crear_archivo_original(
    session: AsyncSession,
    ruta_pdf: Path,
    nombre_archivo: str
) -> Optional[int]:
    """
    Crea o recupera un registro en archivos_originales para el PDF.

    Returns:
        ID del archivo en la BD
    """
    # Calcular metadata
    tamano = ruta_pdf.stat().st_size
    hash_sha256 = calcular_sha256(ruta_pdf)
    paginas = contar_paginas_pdf(ruta_pdf)

    # Verificar si ya existe
    existing = await session.execute(
        text("SELECT id FROM legal.archivos_originales WHERE hash_sha256 = :hash"),
        {"hash": hash_sha256}
    )
    row = existing.fetchone()
    if row:
        return row[0]

    # Crear registro
    ruta_storage = f"legal/guias_sea/{nombre_archivo}"

    result = await session.execute(
        text("""
            INSERT INTO legal.archivos_originales
            (nombre_original, nombre_storage, ruta_storage, mime_type,
             tamano_bytes, hash_sha256, paginas, procesado_at, subido_por)
            VALUES (:nombre_original, :nombre_storage, :ruta_storage,
                    :mime_type, :tamano, :hash, :paginas, :procesado_at, :subido_por)
            RETURNING id
        """),
        {
            "nombre_original": nombre_archivo,
            "nombre_storage": nombre_archivo,
            "ruta_storage": ruta_storage,
            "mime_type": "application/pdf",
            "tamano": tamano,
            "hash": hash_sha256,
            "paginas": paginas,
            "procesado_at": datetime.now(),
            "subido_por": "ingestar_guias_sea.py",
        }
    )
    return result.scalar()


def detectar_triggers(texto: str) -> List[str]:
    """Detecta triggers del Art. 11 en el texto."""
    texto_lower = texto.lower()
    triggers = []

    for letra, keywords in TRIGGERS_KEYWORDS.items():
        for keyword in keywords:
            if keyword.lower() in texto_lower:
                triggers.append(letra)
                break

    return triggers


def detectar_componentes(texto: str) -> List[str]:
    """Detecta componentes ambientales mencionados."""
    componentes_keywords = {
        'agua': ['agua', 'hídrico', 'acuífero', 'río', 'lago', 'humedal'],
        'aire': ['aire', 'emisión', 'atmosférico', 'calidad del aire'],
        'suelo': ['suelo', 'erosión', 'contaminación del suelo'],
        'flora': ['flora', 'vegetación', 'especies vegetales', 'bosque'],
        'fauna': ['fauna', 'especies animales', 'biodiversidad'],
        'glaciares': ['glaciar', 'criósfera', 'permafrost'],
        'patrimonio': ['patrimonio', 'arqueológico', 'monumento'],
        'ruido': ['ruido', 'acústico', 'vibración'],
        'paisaje': ['paisaje', 'visual', 'escénico'],
    }

    texto_lower = texto.lower()
    componentes = []

    for comp, keywords in componentes_keywords.items():
        for keyword in keywords:
            if keyword in texto_lower:
                componentes.append(comp)
                break

    return componentes


async def documento_existe(session: AsyncSession, titulo: str) -> bool:
    """Verifica si un documento ya existe en el corpus."""
    result = await session.execute(
        text("SELECT id FROM legal.documentos WHERE titulo = :titulo"),
        {"titulo": titulo}
    )
    return result.fetchone() is not None


async def ingestar_documento(
    session: AsyncSession,
    doc_data: Dict,
    contenido: str,
    embedding_service,
    ruta_pdf: Path = None,
    nombre_archivo: str = None
) -> Optional[int]:
    """Ingesta un documento al corpus RAG con vinculación de PDF."""

    titulo = doc_data['nombre']

    # Verificar si ya existe
    if await documento_existe(session, titulo):
        logger.info(f"  Ya existe: {titulo[:50]}...")
        return None

    # Detectar triggers y componentes
    triggers = detectar_triggers(contenido)
    componentes = detectar_componentes(contenido)

    # Parsear año como fecha
    fecha_pub = None
    if doc_data.get('anio'):
        try:
            anio = int(str(doc_data['anio']).split('.')[0])
            fecha_pub = datetime(anio, 1, 1).date()
        except:
            pass

    # Crear/obtener archivo original si se proporciona PDF
    archivo_id = None
    if ruta_pdf and nombre_archivo and ruta_pdf.exists():
        archivo_id = await crear_archivo_original(session, ruta_pdf, nombre_archivo)

    # Insertar documento con archivo_id
    result = await session.execute(
        text("""
            INSERT INTO legal.documentos
            (titulo, tipo, fecha_publicacion, organismo, url_fuente,
             contenido_completo, triggers_art11, componentes_ambientales, estado, archivo_id)
            VALUES (:titulo, :tipo, :fecha, :organismo, :url,
                    :contenido, :triggers, :componentes, 'vigente', :archivo_id)
            RETURNING id
        """),
        {
            "titulo": titulo,
            "tipo": doc_data['documento_tipo'],
            "fecha": fecha_pub,
            "organismo": "Servicio de Evaluación Ambiental",
            "url": doc_data['url'],
            "contenido": contenido,
            "triggers": triggers,
            "componentes": componentes,
            "archivo_id": archivo_id,
        }
    )
    documento_id = result.scalar()

    # Segmentar y crear fragmentos
    fragmentos = segmentar_texto(contenido)

    if not fragmentos:
        logger.warning(f"  Sin fragmentos para: {titulo[:50]}")
        return documento_id

    # Generar embeddings en batch
    textos = [f['contenido'] for f in fragmentos]
    embeddings = embedding_service.embed_texts(textos)

    # Insertar fragmentos
    for frag, embedding in zip(fragmentos, embeddings):
        embedding_str = "[" + ",".join(str(v) for v in embedding) + "]"

        await session.execute(
            text(f"""
                INSERT INTO legal.fragmentos
                (documento_id, seccion, numero_seccion, contenido, temas, embedding)
                VALUES (:doc_id, :seccion, :num, :contenido, :temas, '{embedding_str}'::vector)
            """),
            {
                "doc_id": documento_id,
                "seccion": frag['seccion'],
                "num": frag['numero'],
                "contenido": frag['contenido'],
                "temas": frag.get('temas', ['guia_sea']),
            }
        )

    await session.commit()
    logger.info(f"  Ingestado: {titulo[:50]}... ({len(fragmentos)} fragmentos)")

    return documento_id


def segmentar_texto(contenido: str, max_chars: int = 1500) -> List[Dict]:
    """Segmenta texto en fragmentos."""
    import re

    fragmentos = []

    # Intentar segmentar por secciones numeradas
    patron = r'(\d+\.?\d*[\.\-\s]+[A-ZÁÉÍÓÚ][^\.]{10,})'
    secciones = re.split(patron, contenido)

    if len(secciones) > 3:
        seccion_num = 0
        for i in range(0, len(secciones), 2):
            texto = secciones[i]
            if i + 1 < len(secciones):
                texto = secciones[i + 1] + texto

            texto = texto.strip()
            if len(texto) < 100:
                continue

            seccion_num += 1

            # Si es muy largo, dividir
            if len(texto) > max_chars:
                chunks = dividir_en_chunks(texto, max_chars)
                for j, chunk in enumerate(chunks):
                    fragmentos.append({
                        'seccion': f'Sección {seccion_num}' + (f'.{j+1}' if len(chunks) > 1 else ''),
                        'numero': str(seccion_num),
                        'contenido': chunk,
                    })
            else:
                fragmentos.append({
                    'seccion': f'Sección {seccion_num}',
                    'numero': str(seccion_num),
                    'contenido': texto,
                })

    # Si no hay secciones claras, dividir por chunks
    if not fragmentos:
        chunks = dividir_en_chunks(contenido, max_chars)
        for i, chunk in enumerate(chunks):
            fragmentos.append({
                'seccion': f'Fragmento {i+1}',
                'numero': str(i+1),
                'contenido': chunk,
            })

    return fragmentos


def dividir_en_chunks(texto: str, max_chars: int = 1500) -> List[str]:
    """Divide texto largo en chunks respetando oraciones."""
    import re

    if len(texto) <= max_chars:
        return [texto]

    chunks = []
    oraciones = re.split(r'(?<=[.!?])\s+', texto)
    chunk_actual = ""

    for oracion in oraciones:
        if len(chunk_actual) + len(oracion) + 1 <= max_chars:
            chunk_actual += (" " if chunk_actual else "") + oracion
        else:
            if chunk_actual:
                chunks.append(chunk_actual)
            chunk_actual = oracion

    if chunk_actual:
        chunks.append(chunk_actual)

    return chunks


async def main():
    """Proceso principal de ingesta."""
    print("=" * 70)
    print("INGESTA DE GUÍAS Y CRITERIOS DEL SEA")
    print("Sistema de Prefactibilidad Ambiental Minera")
    print("=" * 70)

    # Verificar que existe el archivo Excel
    if not EXCEL_PATH.exists():
        logger.error(f"No se encontró el archivo Excel: {EXCEL_PATH}")
        return

    # Extraer datos del Excel
    guias, criterios = extraer_guias_excel(EXCEL_PATH)
    documentos = guias + criterios

    print(f"\nDocumentos a procesar: {len(documentos)}")
    print(f"  - Guías: {len(guias)}")
    print(f"  - Criterios: {len(criterios)}")
    print()

    # Crear directorio de descarga
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    # Crear conexión a BD
    engine = create_async_engine(DATABASE_URL, echo=False)
    async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    # Importar servicio de embeddings
    from app.services.rag.embeddings import get_embedding_service
    embedding_service = get_embedding_service()

    # Estadísticas
    stats = {
        "descargados": 0,
        "ingestados": 0,
        "existentes": 0,
        "errores": 0,
    }

    async with httpx.AsyncClient() as client:
        async with async_session() as session:
            for i, doc in enumerate(documentos):
                print(f"[{i+1}/{len(documentos)}] {doc['nombre'][:60]}...")

                # Generar nombre de archivo
                nombre_archivo = hashlib.md5(doc['url'].encode()).hexdigest()[:12] + ".pdf"
                ruta_pdf = DOWNLOAD_DIR / nombre_archivo

                # Descargar si no existe
                if not ruta_pdf.exists():
                    ok = await descargar_pdf(doc['url'], ruta_pdf, client)
                    if ok:
                        stats["descargados"] += 1
                    else:
                        stats["errores"] += 1
                        continue

                # Extraer texto
                contenido = extraer_texto_pdf(ruta_pdf)
                if len(contenido) < 200:
                    logger.warning(f"  Contenido muy corto, omitiendo")
                    stats["errores"] += 1
                    continue

                # Ingestar con vinculación de PDF
                try:
                    doc_id = await ingestar_documento(
                        session, doc, contenido, embedding_service,
                        ruta_pdf=ruta_pdf,
                        nombre_archivo=nombre_archivo
                    )
                    if doc_id:
                        stats["ingestados"] += 1
                    else:
                        stats["existentes"] += 1
                except Exception as e:
                    logger.error(f"  Error ingiriendo: {e}")
                    stats["errores"] += 1

    await engine.dispose()

    # Resumen
    print()
    print("=" * 70)
    print("RESUMEN")
    print("=" * 70)
    print(f"PDFs descargados:    {stats['descargados']}")
    print(f"Documentos nuevos:   {stats['ingestados']}")
    print(f"Ya existentes:       {stats['existentes']}")
    print(f"Errores:             {stats['errores']}")
    print()


if __name__ == "__main__":
    asyncio.run(main())
